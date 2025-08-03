import json
import os
import re
import smtplib
import time
from datetime import datetime, timedelta
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from queue import Queue
from threading import Thread, Lock

import openpyxl
from PyQt6 import QtWidgets, QtCore, QtGui
from PyQt6.QtCore import pyqtSignal

from System_info import read_key_value
from Thread import WorkerThread, ErrorSoundThread
from common import get_resource_path, log, str_to_bool, log_print


class AutoInfo(QtWidgets.QWidget):
    update_ui_signal = pyqtSignal(int, str, str, str, list, str)

    def __init__(self, wx_dict, membership, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.wx_dict = wx_dict
        self.membership = membership

        self.ready_tasks = {}
        self.completed_tasks = {}
        self.task_id_counter = 0
        self.tasks_by_time = {}

        self.weekday_map = {
            "仅一次": [],
            "星期一": [1],
            "星期二": [2],
            "星期三": [3],
            "星期四": [4],
            "星期五": [5],
            "星期六": [6],
            "星期日": [7],
            "每天": list(range(1, 8)),
            "工作日": list(range(1, 6)),
            "周末末": [6, 7]
        }

        self.reverse_weekday_map = {
            1: "星期一",
            2: "星期二",
            3: "星期三",
            4: "星期四",
            5: "星期五",
            6: "星期六",
            7: "星期日"
        }

        self.is_executing = False
        self.worker_thread = None
        self.error_sound_thread = ErrorSoundThread()
        self.audio_files = {
            0: get_resource_path('resources/sound/error_sound_1.mp3'),
            1: get_resource_path('resources/sound/error_sound_2.mp3'),
            2: get_resource_path('resources/sound/error_sound_3.mp3'),
            3: get_resource_path('resources/sound/error_sound_4.mp3'),
            4: get_resource_path('resources/sound/error_sound_5.mp3')
        }
        self.email_queue = Queue()
        self.email_thread = Thread(target=self.process_email_queue, daemon=True)
        self.email_thread.start()
        self.last_email_time = 0
        self.email_cooldown = 60

        self.save_lock = Lock()
        self.save_pending = False

        self.update_ui_signal.connect(self.add_task_to_ui)
        log_print(f"[AutoInfo] Initialization completed. Membership level: {self.membership}")

    def get_unique_task_id(self, preferred_id=None):
        if preferred_id is not None and preferred_id not in self.ready_tasks and preferred_id not in self.completed_tasks:
            self.task_id_counter = max(self.task_id_counter, preferred_id)
            return preferred_id

        self.task_id_counter += 1
        return self.task_id_counter

    def save_tasks_to_json(self):
        try:
            log_print("[AutoInfo] Saving tasks to JSON file...")
            tasks_list = []
            for task in self.ready_tasks.values():
                tasks_data = {
                    'id': task['id'],
                    'time': task['time'],
                    'sender': task['sender'],
                    'name': task['name'],
                    'info': task['info'],
                    'frequency': task['frequency']
                }
                tasks_list.append(tasks_data)

            for task in self.completed_tasks.values():
                task_data = {
                    'id': task['id'],
                    'time': task['time'],
                    'sender': task['sender'],
                    'name': task['name'],
                    'info': task['info'],
                    'frequency': task['frequency'],
                    'status': task.get('status', '')
                }
                tasks_list.append(task_data)

            log_print(f"[AutoInfo] Preparing to save {len(tasks_list)} tasks to JSON file...")

            if not os.path.exists('_internal'):
                os.makedirs('_internal')

            with open('_internal/tasks.json', 'w', encoding='utf-8') as f:
                json.dump(tasks_list, f, ensure_ascii=False, indent=4)
        except Exception as e:
            log_print(f"[AutoInfo] Failed to save tasks: {str(e)}")
            log("ERROR", "非管理员身份运行软件,未能将操作保存")

    def delayed_save(self):
        log_print("[AutoInfo] Preparing for delayed task saving...")
        with self.save_lock:
            if self.save_pending:
                log_print("[AutoInfo] There is already a save request waiting, canceling current request")
                return
            self.save_pending = True
        Thread(target=self.delayed_save_thread, daemon=True).start()

    def delayed_save_thread(self):
        time.sleep(2)
        self.save_tasks_to_json()
        with self.save_lock:
            self.save_pending = False
        log_print("[AutoInfo] Delayed save thread completed")

    def add_list_item(self):
        try:
            time_text, name_text, info_text, frequency, sender_text = self.get_input_values()

            if not all([time_text, name_text, info_text, sender_text]):
                log("WARNING", "任务添加失败: 输入不完整，缺少必要信息")
                log_print("[AutoInfo] Task addition failed: Incomplete input, missing necessary information")
                return

            if self.membership == 'Free' and len(self.ready_tasks) >= 5:
                log("WARNING", "试用版最多添加5个任务，请升级版本")
                log_print("[AutoInfo] Trial version can add up to 5 tasks, please upgrade upgrade version")
                QtWidgets.QMessageBox.warning(self, "试用版限制", "试用试用版最多添加5个任务，请升级版本")
                return
            elif self.membership == 'Base' and len(self.ready_tasks) >= 30:
                log("WARNING", "任务添加失败: 基础会员限制(30个任务)已达上限")
                log_print("[AutoInfo] Task addition failed: Basic membership limit (30 tasks) reached")
                return

            task_id = self.get_unique_task_id()

            task_data = {
                'id': task_id,
                'time': time_text,
                'sender': sender_text,
                'name': name_text,
                'info': info_text,
                'frequency': frequency
            }

            self.ready_tasks[task_id] = task_data
            if time_text not in self.tasks_by_time:
                self.tasks_by_time[time_text] = []
            self.tasks_by_time[time_text].append(task_id)
            widget_item = self.create_widget(task_id, time_text, name_text, info_text, frequency, sender_text)
            self.parent.formLayout_3.addRow(widget_item)
            log('INFO',
                f'已添加 {time_text[-8:]} 由 {sender_text[:8]} 发送给 {name_text[:8]}: {info_text[:25] + "……" if len(info_text) > 25 else info_text}')
            log_print(
                f"[AutoInfo] Added task {time_text[-8:]} from {sender_text[:8]} to {name_text[:8]}: {info_text[:25] + '...' if len(info_text) > 25 else info_text}")

            timestep = int(read_key_value('add_timestep'))
            self.parent.dateTimeEdit.setDateTime(
                datetime.fromisoformat(time_text) + timedelta(minutes=timestep))
            log_print(f"[AutoInfo] DateTime widget updated, increased by {timestep} minutes")

            self.delayed_save()

        except Exception as e:
            log("ERROR", f"添加任务时发生异常: {str(e)}")
            log_print(f"[AutoInfo] Exception occurred while adding task: {str(e)}")

    def create_widget(self, task_id, time_text, name_text, info_text, frequency, sender_text):
        try:
            log_print(f"[AutoInfo] Creating widget for task: {sender_text} -> {name_text} - {info_text}...")
            widget_item = QtWidgets.QWidget(parent=self.parent.scrollAreaWidgetContents_3)
            widget_item.setMinimumSize(QtCore.QSize(360, 80))
            widget_item.setMaximumSize(QtCore.QSize(360, 80))
            widget_item.setStyleSheet("background-color: rgb(255, 255, 255);\nborder-radius:18px")
            widget_item.setObjectName("widget_item")

            horizontal_layout_76 = QtWidgets.QHBoxLayout(widget_item)
            horizontal_layout_76.setContentsMargins(12, 2, 12, 2)
            horizontal_layout_76.setSpacing(6)
            horizontal_layout_76.setObjectName("horizontalLayout_76")

            widget_54 = QtWidgets.QWidget(parent=widget_item)
            widget_54.setMinimumSize(QtCore.QSize(36, 36))
            widget_54.setMaximumSize(QtCore.QSize(36, 36))
            widget_54.setStyleSheet(f"image: url({get_resource_path('resources/img/page1/page1_发送就绪.svg')});")
            widget_54.setObjectName("widget_54")
            horizontal_layout_76.addWidget(widget_54)

            vertical_layout_64 = QtWidgets.QVBoxLayout()
            vertical_layout_64.setContentsMargins(6, 6, 6, 6)
            vertical_layout_64.setSpacing(0)
            vertical_layout_64.setObjectName("verticalLayout_64")

            horizontal_layout_77 = QtWidgets.QHBoxLayout()
            horizontal_layout_77.setContentsMargins(0, 0, 0, 0)
            horizontal_layout_77.setSpacing(4)
            horizontal_layout_77.setObjectName("horizontalLayout_77")

            sender_receiver_label = QtWidgets.QLabel(f"{sender_text} → {name_text}", parent=widget_item)
            font = QtGui.QFont()
            font.setFamily("微软雅黑 Light")
            font.setPointSize(12)
            sender_receiver_label.setFont(font)
            sender_receiver_label.setStyleSheet("color:rgb(0, 0, 0);")
            sender_receiver_label.setObjectName("sender_receiver_label")
            horizontal_layout_77.addWidget(sender_receiver_label)

            time_label = QtWidgets.QLabel(time_text, parent=widget_item)
            font = QtGui.QFont()
            font.setPointSize(10)
            time_label.setFont(font)
            time_label.setStyleSheet("color: rgb(169, 169, 169);")
            time_label.setAlignment(
                QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignTrailing | QtCore.Qt.AlignmentFlag.AlignVCenter)
            time_label.setObjectName("time_label")
            horizontal_layout_77.addWidget(time_label)

            if not frequency:
                frequency_text = "仅一次"
            else:
                sorted_weekdays = sorted(frequency)
                frequency_text = ",".join([str(num) for num in sorted_weekdays])

            frequency_label = QtWidgets.QLabel(frequency_text, parent=widget_item)
            frequency_label.setStyleSheet("color:rgb(105, 27, 253);\ntext-align: center;\nbackground:rgba(0, 0, 0, 0);")
            frequency_label.setObjectName("label_2")
            horizontal_layout_77.addWidget(frequency_label)

            horizontal_layout_77.setStretch(0, 1)
            vertical_layout_64.addLayout(horizontal_layout_77)

            horizontal_layout_7 = QtWidgets.QHBoxLayout()
            horizontal_layout_7.setContentsMargins(0, 6, 12, 3)
            horizontal_layout_7.setObjectName("horizontalLayout_7")

            message_label = QtWidgets.QLabel(info_text, parent=widget_item)
            font = QtGui.QFont()
            font.setPointSize(10)
            message_label.setFont(font)
            message_label.setStyleSheet("color: rgb(169, 169, 169);")
            message_label.setObjectName("message_label")
            horizontal_layout_7.addWidget(message_label)

            delete_button = QtWidgets.QPushButton("删除", parent=widget_item)
            delete_button.setFixedSize(50, 25)
            delete_button.setStyleSheet(
                "QPushButton { background-color: transparent; color: red; } QPushButton:hover { background-color: rgba(255, 0, 0, 0.1); }"
            )
            delete_button.clicked.connect(
                lambda checked, tid=task_id: self.remove_task(tid))
            delete_button.setVisible(False)

            horizontal_layout_7.addWidget(delete_button)

            widget_item.enterEvent = lambda event, btn=delete_button: btn.setVisible(True)
            widget_item.leaveEvent = lambda event, btn=delete_button: btn.setVisible(False)

            vertical_layout_64.addLayout(horizontal_layout_7)
            horizontal_layout_76.addLayout(vertical_layout_64)

            widget_item.task_id = task_id
            return widget_item

        except Exception as e:
            log_print(f"[AutoInfo] Failed to create task widget: {str(e)}")
            return None

    def remove_task(self, task_id):
        log_print(f"[AutoInfo] Attempting to remove task (ID: {task_id})...")
        try:
            widget = None
            for i in range(self.parent.formLayout_3.count()):
                item = self.parent.formLayout_3.itemAt(i)
                if item and item.widget() and hasattr(item.widget(), 'task_id') and item.widget().task_id == task_id:
                    widget = item.widget()
                    break

            if widget:
                widget.hide()
                QtCore.QTimer.singleShot(0, widget.deleteLater)
                log_print(f"[AutoInfo] Task widget removed from interface (ID: {task_id})")

            def process_data_removal():
                if self.error_sound_thread._is_running:
                    self.error_sound_thread.stop_playback()
                    log_print("[AutoInfo] Error sound stopped")

                if task_id in self.ready_tasks:
                    task = self.ready_tasks.pop(task_id)
                elif task_id in self.completed_tasks:
                    task = self.completed_tasks.pop(task_id)
                else:
                    log_print(f"[AutoInfo] Task data not found (ID: {task_id})")
                    return

                if task:
                    time_text = task['time']
                    if time_text in self.tasks_by_time:
                        try:
                            self.tasks_by_time[time_text].remove(task_id)
                            if not self.tasks_by_time[time_text]:
                                del self.tasks_by_time[time_text]
                                log("INFO", f"时间索引 {time_text} 已删除，因为已无任务")
                                log_print(f"[AutoInfo] Time index {time_text} deleted as no tasks remain")
                        except ValueError:
                            log("WARNING", f"任务ID {task_id} 不在时间索引 {time_text} 中")
                            log_print(f"[AutoInfo] Task ID {task_id} not in time index {time_text}")

                    log('WARNING', f'已删除任务 {task["info"][:35] + "……" if len(task["info"]) > 30 else task["info"]}')
                    log_print(f"[AutoInfo] Task removed: {task['info'][:35] + '...' if len(task['info']) > 30 else task['info']}")

                    if not self.ready_tasks and self.is_executing:
                        self.is_executing = False
                        if self.parent and hasattr(self.parent, 'start_pushButton') and self.parent.start_pushButton:
                            def update_button_text():
                                if self.parent and self.parent.start_pushButton:
                                    self.parent.start_pushButton.setText("开始执行")

                            if QtCore.QThread.currentThread() != self.parent.thread():
                                QtCore.QTimer.singleShot(0, update_button_text)
                            else:
                                update_button_text()

                        if self.worker_thread is not None:
                            self.worker_thread.requestInterruption()
                            self.worker_thread = None
                            log("INFO", "工作线程已停止")
                            log_print("[AutoInfo] Worker thread stopped")

                    self.delayed_save()

            QtCore.QTimer.singleShot(100, process_data_removal)

        except Exception as e:
            log("ERROR", f"删除任务失败: {str(e)}")
            log_print(f"[AutoInfo] Failed to remove task: {str(e)}")

    def update_ui(self):
        try:
            self.clear_layout(self.parent.formLayout_3)

            sorted_task_ids = sorted(self.ready_tasks.keys(),
                                     key=lambda tid: self.ready_tasks[tid]['time'])
            for task_id in sorted_task_ids:
                task = self.ready_tasks[task_id]
                widget = self.create_widget(
                    task_id, task['time'], task['name'], task['info'], task['frequency'], task['sender'])
                self.parent.formLayout_3.addRow(widget)
            log_print("[AutoInfo] Interface updated successfully")

        except Exception as e:
            log_print(f"[AutoInfo] Failed to update interface: {str(e)}")

    def get_input_values(self):
        try:
            sender_text = self.parent.comboBox_nickName.currentText()
            name_text = self.parent.receiver_lineEdit.text()
            info_text = self.parent.message_lineEdit.text()
            time_text = self.parent.dateTimeEdit.dateTime().toString(QtCore.Qt.DateFormat.ISODate)

            if hasattr(self.parent.comboBox_Frequency, 'get_checked_items'):
                checked_items = self.parent.comboBox_Frequency.get_checked_items()
                frequency = []
                for item in checked_items:
                    if item in self.weekday_map:
                        frequency.extend(self.weekday_map[item])
                frequency = list(set(frequency))
            else:
                frequency_text = self.parent.comboBox_Frequency.currentText()
                frequency = self.weekday_map.get(frequency_text, [])
            log_print(
                f"[AutoInfo] Input values obtained successfully: Sender={sender_text}, Receiver={name_text}, Time={time_text}, Frequency={frequency}")
            return time_text, name_text, info_text, frequency, sender_text

        except Exception as e:
            log_print(f"[AutoInfo] Failed to get input values: {str(e)}")
            return None, None, None, None, None

    def on_start_clicked(self):
        if self.is_executing:
            self.is_executing = False
            self.parent.start_pushButton.setText("开始执行")
            if self.worker_thread is not None:
                self.worker_thread.requestInterruption()
                self.worker_thread = None
                log("INFO", "工作线程已中断")
                log_print("[AutoInfo] Worker thread interrupted")
            if self.error_sound_thread._is_running:
                self.error_sound_thread.stop_playback()
                log_print("[AutoInfo] Error sound stopped")
        else:
            if not self.ready_tasks:
                log("WARNING", "任务列表为空，请先添加任务至任务列表")
                log_print("[AutoInfo] Task list is empty, please add tasks first")
                return

            current_time = datetime.now()
            has_past_tasks = any(
                datetime.fromisoformat(time_str) < current_time
                for time_str in self.tasks_by_time
            )

            if has_past_tasks:
                message = (
                    f"<p style='color:#d9534f;font-size:16px;'>发现已过期的定时任务</p>"
                    f"<p>点击确定将立即执行这些任务，否则请重新设置再启动</p>"
                )

                msg_box = QtWidgets.QMessageBox(
                    QtWidgets.QMessageBox.Icon.Warning,
                    "过期任务二次确认",
                    message,
                    QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                    self
                )
                msg_box.setDefaultButton(QtWidgets.QMessageBox.StandardButton.No)
                reply = msg_box.exec()

                if reply == QtWidgets.QMessageBox.StandardButton.No:
                    log_print("[AutoInfo] User canceled execution of expired tasks")
                    return

            self.is_executing = True
            self.parent.start_pushButton.setText("停止执行")
            self.worker_thread = WorkerThread(self)
            self.worker_thread.prevent_sleep = self.parent.checkBox_stopSleep.isChecked()
            self.worker_thread.current_time = 'mix' if str_to_bool(read_key_value('net_time')) else 'sys'
            self.worker_thread.finished.connect(self.on_thread_finished)
            self.worker_thread.start()

    def clear_layout(self, layout):
        try:
            log_print("[AutoInfo] Clearing layout...")
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
        except Exception as e:
            log_print(f"[AutoInfo] Failed to clear layout: {str(e)}")

    def update_task_status(self, task_id, status):
        try:
            if task_id not in self.ready_tasks:
                log_print(f"[AutoInfo] Task ID {task_id} not found for status update")
                return

            task = self.ready_tasks.pop(task_id)
            time_text = task['time']

            if time_text in self.tasks_by_time:
                try:
                    self.tasks_by_time[time_text].remove(task_id)
                    if not self.tasks_by_time[time_text]:
                        del self.tasks_by_time[time_text]
                        log_print(f"[AutoInfo] Time index {time_text} deleted as no tasks remain")
                except ValueError:
                    log_print(f"[AutoInfo] Task ID {task_id} not in time index {time_text}")

            task['status'] = status
            self.completed_tasks[task_id] = task
            for i in range(self.parent.formLayout_3.count()):
                item = self.parent.formLayout_3.itemAt(i)
                if item and item.widget() and hasattr(item.widget(), 'task_id') and item.widget().task_id == task_id:
                    widget_item = item.widget()
                    widget_image = widget_item.findChild(QtWidgets.QWidget, "widget_54")
                    if widget_image:
                        if status == '成功':
                            icon_path = 'page1_发送成功.svg'
                        else:
                            icon_path = 'page1_发送失败.svg'

                            if 'error_count' not in task:
                                task['error_count'] = 1
                                self.play_error_sound()
                                self.send_error_email(task)
                            else:
                                task['error_count'] += 1
                                log("WARNING",
                                    f"任务 {task['sender']} -> {task['name']} 已失败 {task['error_count']} 次")
                                log_print(
                                    f"[AutoInfo] Task {task['sender']} -> {task['name']} has failed {task['error_count']} times")

                        new_icon_path = get_resource_path(f'resources/img/page1/{icon_path}')
                        widget_image.setStyleSheet(f"image: url({new_icon_path});")
                        log_print(f"[AutoInfo] Task icon updated to {icon_path} (ID: {task_id})")

                    if task['frequency']:
                        current_time = datetime.fromisoformat(task['time'])
                        next_time = self.calculate_next_time(current_time, task['frequency'])
                        if next_time:
                            self.update_ui_signal.emit(task_id, next_time.isoformat(), task['name'], task['info'],
                                                       task['frequency'], task['sender'])
                    self.delayed_save()
                    break

        except Exception as e:
            log_print(f"[AutoInfo] Failed to update task status: {str(e)}")

    def calculate_next_time(self, current_time, frequency):
        if not frequency:
            log_print("[AutoInfo] Frequency is empty, not calculating next time")
            return None

        now = datetime.now()
        start_time = current_time if current_time > now else now
        log_print(f"[AutoInfo] Starting calculation of next task time, start time: {start_time}")

        next_time = start_time
        max_days = 30
        days_checked = 0

        while days_checked < max_days:
            weekday_num = next_time.weekday() + 1

            if weekday_num in frequency:
                candidate_time = next_time.replace(
                    hour=current_time.hour,
                    minute=current_time.minute,
                    second=current_time.second,
                    microsecond=0
                )

                if candidate_time > now:
                    log_print(f"[AutoInfo] Calculated next task time: {candidate_time}")
                    return candidate_time

            next_time += timedelta(days=1)
            days_checked += 1
        log_print("[AutoInfo] Exceeded maximum check days, no suitable next time found")
        return None

    def add_task_to_ui(self, original_task_id, time_text, name_text, info_text, frequency, sender_text):
        try:
            task_id = self.get_unique_task_id()
            task_data = {
                'id': task_id,
                'time': time_text,
                'sender': sender_text,
                'name': name_text,
                'info': info_text,
                'frequency': frequency
            }
            self.ready_tasks[task_id] = task_data
            log_print(f"[AutoInfo] Added recurring task data (ID: {task_id})")

            if time_text not in self.tasks_by_time:
                self.tasks_by_time[time_text] = []
            self.tasks_by_time[time_text].append(task_id)
            log_print(f"[AutoInfo] Task added to time index: {time_text} (ID: {task_id})")

            widget_item = self.create_widget(task_id, time_text, name_text, info_text, frequency, sender_text)
            if widget_item:
                sorted_task_ids = sorted(self.ready_tasks.keys(),
                                         key=lambda tid: self.ready_tasks[tid]['time'])
                task_index = sorted_task_ids.index(task_id)
                self.parent.formLayout_3.insertRow(task_index, widget_item)

                self.parent.formLayout_3.update()
                self.parent.scrollAreaWidgetContents_3.updateGeometry()
                self.parent.scrollArea_3.update()

                log('INFO',
                    f'自动添加 {time_text} 由 {sender_text[:8]} 发送给 {name_text[:8]}: {info_text[:25] + "……" if len(info_text) > 25 else info_text}')
                log_print(
                    f"[AutoInfo] Automatically added {time_text} from {sender_text[:8]} to {name_text[:8]}: {info_text[:25] + '...' if len(info_text) > 25 else info_text}")
                self.delayed_save()
            else:
                self.task_id_counter -= 1
                if task_id in self.ready_tasks:
                    del self.ready_tasks[task_id]
                if time_text in self.tasks_by_time and task_id in self.tasks_by_time[time_text]:
                    self.tasks_by_time[time_text].remove(task_id)
                log_print(f"[AutoInfo] Failed to create recurring task widget, rolled back operation (ID: {task_id})")

        except Exception as e:
            log_print(f"[AutoInfo] Failed to add recurring task to interface: {str(e)}")
            self.task_id_counter -= 1

    def on_thread_finished(self):
        try:
            self.is_executing = False
            self.parent.start_pushButton.setText("开始执行")

            if self.parent.checkBox_Shutdown.isChecked():
                self.shutdown_computer()

        except Exception as e:
            log("ERROR", f"处理线程完成事件失败: {str(e)}")
            log_print(f"[AutoInfo] Failed to process thread completion event: {str(e)}")

    def shutdown_computer(self):
        log("INFO", "准备关闭计算机...")
        log_print("[AutoInfo] Preparing to shut down computer...")

        try:
            for i in range(10, 0, -1):
                log("WARNING", f"电脑在 {i} 秒后自动关机")
                log_print(f"[AutoInfo] Computer will shut down automatically in {i} seconds")
                time.sleep(1)

            log("DEBUG", "正在关机中...")
            log_print("[AutoInfo] Shutting down...")
            os.system('shutdown /s /t 0')

        except Exception as e:
            log("ERROR", f"关闭计算机失败: {str(e)}")
            log_print(f"[AutoInfo] Failed to shut down computer: {str(e)}")

    def save_configuration(self):

        try:
            if not self.ready_tasks:
                log("WARNING", "当前任务列表为空,没有任务可供保存")
                log_print("[AutoInfo] Current task list is empty, no tasks to save")
                return

            current_date = datetime.now().strftime("%m%d")
            default_filename = f"LeafAuto自动计划_{current_date}"

            documents_dir = os.path.expanduser("~/Documents")
            file_name, _ = QtWidgets.QFileDialog.getSaveFileName(
                self,
                "保存任务计划",
                os.path.join(documents_dir, default_filename),
                "枫叶任务文件(*.xlsx);;所有文件(*)"
            )

            if file_name:
                if not file_name.lower().endswith('.xlsx'):
                    file_name += '.xlsx'

                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(['ID', 'Time', 'Sender', 'Name', 'Info', 'Frequency'])

                sorted_tasks = sorted(self.ready_tasks.values(), key=lambda x: x['time'])
                for task in sorted_tasks:
                    if not task['frequency']:
                        freq_text = "仅一次"
                    else:
                        freq_text = ",".join(map(str, task['frequency']))
                    sheet.append([task['id'], task['time'], task['sender'], task['name'], task['info'], freq_text])

                workbook.save(file_name)
        except Exception as e:
            log("ERROR", f"保存配置失败: {str(e)}")
            log_print(f"[AutoInfo] Failed to save configuration: {str(e)}")

    def open_file_dialog(self, filepath=None):
        try:
            if filepath:
                self.parent.message_lineEdit.setText(str(filepath))
                log_print(f"[AutoInfo] File path set: {filepath}")
                return

            file_filters = (
                "所有文件 (*);;"
                "图像文件 (*.bmp *.gif *.jpg *.jpeg *.png *.svg *.tiff);;"
                "文档文件 (*.doc *.docx *.pdf *.txt *.odt);;"
                "电子表格 (*.xls *.xlsx *.ods);;"
                "演示文稿 (*.ppt *.pptx *.odp);;"
                "音频文件 (*.mp3 *.wav *.flac *.aac);;"
                "视频文件 (*.mp4 *.avi *.mkv *.mov);;"
                "压缩文件 (*.zip *.rar *.tar *.gz *.bz2)"
            )

            options = QtWidgets.QFileDialog.Option.ReadOnly
            file_name, _ = QtWidgets.QFileDialog.getOpenFileName(
                None,
                "选择要发送的文件",
                "",
                file_filters,
                options=options
            )

            if file_name:
                self.parent.message_lineEdit.setText(file_name)
                log_print(f"[AutoInfo] File selected: {file_name}")
        except Exception as e:
            log_print(f"[AutoInfo] Failed to open file dialog: {str(e)}")

    def load_configuration(self, filepath=None):
        filter_expired = str_to_bool(read_key_value("import_filter"))
        expired_count = 0

        try:
            documents_dir = os.path.expanduser("~/Documents")
            file_name = filepath or QtWidgets.QFileDialog.getOpenFileName(
                self, "导入任务计划", documents_dir, "枫叶任务文件(*.xlsx);;所有文件(*)"
            )[0]

            if not file_name:
                return

            workbook = openpyxl.load_workbook(file_name, read_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            required_headers = ['Time', 'Sender', 'Name', 'Info', 'Frequency']
            if not all(h in headers for h in required_headers):
                raise ValueError("文件格式不正确，缺少必要的列（可能缺少发送方列）")

            has_id_column = 'ID' in headers

            tasks = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                task = dict(zip(headers, row))

                if not (task.get('Time') and task.get('Sender') and task.get('Name') and task.get('Info')):
                    log("WARNING", f"行 {row_idx} 缺少必要字段，已跳过")
                    log_print(f"[AutoInfo] Row {row_idx} missing necessary fields, skipped")
                    continue

                if not re.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}$", task['Time']):
                    log("WARNING", f"行 {row_idx} 时间格式不正确，已跳过")
                    log_print(f"[AutoInfo] Row {row_idx} has incorrect time format, skipped")
                    continue

                try:
                    task_time = datetime.fromisoformat(task['Time'])
                except Exception as e:
                    log("WARNING", f"行 {row_idx} 时间解析失败，跳过任务: {str(e)}")
                    log_print(f"[AutoInfo] Row {row_idx} time parsing failed, skipping task: {str(e)}")
                    continue

                freq_text = task.get('Frequency', "仅一次")
                if freq_text == "仅一次":
                    task['Frequency'] = []
                else:
                    freq_list = []
                    for day in freq_text.split(','):
                        if day.strip().isdigit():
                            num = int(day.strip())
                            if 1 <= num <= 7:
                                freq_list.append(num)
                        else:
                            for num, name in self.reverse_weekday_map.items():
                                if name == day.strip():
                                    freq_list.append(num)
                                    break
                    task['Frequency'] = list(set(freq_list))
                    freq_list = task['Frequency']

                current_time = datetime.now()
                if task_time < current_time:
                    if not filter_expired and task['Frequency']:
                        adjusted_time = self.calculate_next_time(task_time, task['Frequency'])
                        if adjusted_time:
                            task['Time'] = adjusted_time.isoformat()
                            log_print(f"[AutoInfo] Row {row_idx} task time adjusted to future time: {task['Time']}")
                        else:
                            log_print(f"[AutoInfo] Row {row_idx} cannot adjust task time, using original time")
                    elif filter_expired:
                        expired_count += 1
                        continue

                task['Sender'] = str(task['Sender']) if task['Sender'] is not None else ""
                task['Name'] = str(task['Name']) if task['Name'] is not None else ""
                task['Info'] = str(task['Info']) if task['Info'] is not None else ""
                tasks.append(task)
                log_print(f"[AutoInfo] Row {row_idx} task parsed successfully")

            workbook.close()

            if filter_expired and expired_count > 0:
                QtWidgets.QMessageBox.information(
                    self,
                    "过滤过期完成",
                    f"已为您过滤{expired_count}个过期任务"
                )
                log("INFO", f"已过滤 {expired_count} 个过期任务")
                log_print(f"[AutoInfo] Filtered out {expired_count} expired tasks")

        except Exception as e:
            log("ERROR", f"导入失败: {str(e)}")
            log_print(f"[AutoInfo] Import failed: {str(e)}")
            return

        if not tasks:
            log("ERROR", f"导入失败, 未找到有效任务数据")
            log_print("[AutoInfo] Import failed, no valid task data found")
            return

        membership_limits = {
            'Free': 5,
            'Base': 30,
            'AiVIP': float('inf'),
            'VIP': float('inf')
        }

        limit = membership_limits.get(self.membership, 5)
        current_count = len(self.ready_tasks)
        if current_count + len(tasks) > limit:
            remaining = limit - current_count
            if remaining <= 0:
                log("WARNING", f"会员限制, 当前版本最多支持{limit}个任务")
                log_print(f"[AutoInfo] Membership limit, current version supports up to {limit} tasks")
                return

            tasks = tasks[:remaining]
            log_print(f"[AutoInfo] Due to membership restrictions, only first {remaining} tasks imported")

        existing_ids = set(self.ready_tasks.keys()).union(set(self.completed_tasks.keys()))

        if existing_ids:
            max_existing_id = max(existing_ids)
            if self.task_id_counter <= max_existing_id:
                self.task_id_counter = max_existing_id
                log_print(f"[AutoInfo] Task ID counter updated to maximum existing ID: {self.task_id_counter}")

        for task in tasks:
            try:
                file_id = int(task['ID']) if has_id_column and task.get('ID') is not None and str(
                    task['ID']).isdigit() else None
                task_id = self.get_unique_task_id(preferred_id=file_id)

                task_data = {
                    'id': task_id,
                    'time': task['Time'],
                    'sender': task['Sender'],
                    'name': task['Name'],
                    'info': task['Info'],
                    'frequency': task['Frequency']
                }

                self.ready_tasks[task_id] = task_data

                time_text = task['Time']
                if time_text not in self.tasks_by_time:
                    self.tasks_by_time[time_text] = []
                self.tasks_by_time[time_text].append(task_id)

                widget = self.create_widget(task_id, task['Time'], task['Name'], task['Info'], task['Frequency'],
                                            task['Sender'])
                self.parent.formLayout_3.addRow(widget)
            except Exception as e:
                log("ERROR", f"导入任务失败: {task['Sender']} -> {task['Name']}, 错误: {str(e)}")
                log_print(f"[AutoInfo] Failed to import task: {task['Sender']} -> {task['Name']}, error: {str(e)}")

        log("INFO", f"成功导入 {len(tasks)} 个任务")
        log_print(f"[AutoInfo] Successfully imported {len(tasks)} tasks")
        self.delayed_save()

    def add_emotion_to_message(self):
        log_print("[AutoInfo] Adding emotion to message")

    def play_error_sound(self):
        log_print("[AutoInfo] Attempting to play error sound...")

        try:
            if str_to_bool(read_key_value('error_sound')):
                if self.error_sound_thread._is_running:
                    log_print("[AutoInfo] Error sound is already playing, skipped")
                    return

                try:
                    selected_audio_index = int(read_key_value('selected_audio_index'))
                except Exception:
                    selected_audio_index = 0
                    log_print("[AutoInfo] Audio index retrieval failed, using default value 0")

                if selected_audio_index in self.audio_files:
                    self.selected_audio_file = self.audio_files[selected_audio_index]
                else:
                    log_print(f"[AutoInfo] Audio playback failed: invalid index {selected_audio_index}")
                    return

                self.error_sound_thread.update_sound_file(self.selected_audio_file)
                self.error_sound_thread.start()
                log_print(f"[AutoInfo] Error sound playing: {self.selected_audio_file}")

        except Exception as e:
            log_print(f"[AutoInfo] Failed to play error sound: {str(e)}")

    def send_error_email(self, task):

        if str_to_bool(read_key_value('error_email')):
            self.email_queue.put(task)

    def process_email_queue(self):
        while True:
            task = self.email_queue.get()
            if task is None:
                break
            self.send_email_safely(task)
            self.email_queue.task_done()

    def send_email_safely(self, task):
        try:
            current_time = time.time()
            if current_time - self.last_email_time < self.email_cooldown:
                return

            log_print(f"[AutoInfo] Sending task error email: {task['sender']} -> {task['name']}")

            sender_email = '3555844679@qq.com'
            receiver_email = read_key_value('email')
            smtp_server = 'smtp.qq.com'
            smtp_port = 465
            username = '3555844679@qq.com'
            password = 'xtibpzrdwnppchhi'

            if not task['frequency']:
                freq_display = "仅一次"
            else:
                freq_display = ",".join([self.reverse_weekday_map[num] for num in task['frequency']])

            subject = f"定时信息发送失败 {task['time']}"

            message = MIMEMultipart('related')
            message['From'] = 'LeafAuto <3555844679@qq.com>'
            message['To'] = receiver_email
            message['Subject'] = Header(subject, 'utf-8')

            html_content = f"""
                    <html>
                    <head>
                        <meta charset="utf-8">
                        <style>
                            body {{ font-family: Arial, sans-serif; margin: 0; padding: 0; background-color: #f8f9fa; }}
                            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                            .header {{ background-color: #4a6fa5; color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
                            .content {{ background-color: white; padding: 20px; border-radius: 0 0 5px 5px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
                            .section {{ margin-bottom: 20px; }}
                            .highlight {{ background-color: #f0f4f8; padding: 10px; border-radius: 3px; }}
                            .button {{ display: inline-block; background-color: #4a6fa5; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; }}
                            .footer {{ margin-top: 20px; text-align: center; color: #6c757d; font-size: 12px; }}
                            .support {{ margin-top: 10px; }}
                            .separator {{ border-top: 1px solid #e9ecef; margin: 20px 0; }}
                            .footer-image {{ max-width: 100%; margin-top: 20px; border-radius: 3px; }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <h2>LeafAuto 消息通知</h2>
                            </div>
                            <div class="content">
                                <div class="section">
                                    <p>尊敬的用户,</p>
                                    <p style="color: #d9534f; font-weight: bold;">😔 非常抱歉地通知您，枫叶未能成功发送您设置的定时信息。</p>
                                </div>

                                <div class="section highlight">
                                    <h3>📌 失败任务详情</h3>
                                    <p>🕒 时间: {task['time']}</p>
                                    <p>✉️ 发送方: {task['sender']}</p>
                                    <p>👤 接收人: {task['name']}</p>
                                    <p>💬 消息内容: {task['info']}</p>
                                    <p>🔄 频率: {freq_display}</p>
                                    <p style="color: #d9534f;">❌ 失败原因: 未知错误。</p>
                                </div>

                                <div class="section">
                                    <h3>🔧 建议操作</h3>
                                    <ol>
                                        <li>检查接收人名称是否与联系人完全一致</li>
                                        <li>确认接收人当前状态是否可以接收消息</li>
                                        <li>尝试手动发送相同内容进行验证</li>
                                    </ol>
                                </div>

                                <div class="section separator"></div>

                                <div class="section">
                                    <p>Dear Valued User 👋,</p>
                                    <p style="color: #d9534f; font-weight: bold;">😔 We sincerely apologize for the inconvenience caused. Our system failed to deliver a scheduled message.</p>
                                </div>

                                <div class="section highlight">
                                    <h3>📌 Delivery Failure Details</h3>
                                    <p>🕒 Time: {task['time']}</p>
                                    <p>✉️ Sender: {task['sender']}</p>
                                    <p>👤 Recipient: {task['name']}</p>
                                    <p>💬 Message: {task['info']}</p>
                                    <p>🔄 Frequency: {freq_display}</p>
                                    <p style="color: #d9534f;">❌ Failure Reason: System error</p>
                                </div>

                                <div class="section">
                                    <h3>🔧 Recommended Actions</h3>
                                    <ol>
                                        <li>Verify the recipient's name matches exactly as in your contacts</li>
                                        <li>Confirm the recipient is currently available to receive messages</li>
                                        <li>Attempt to send the same content manually for validation</li>
                                    </ol>
                                </div>

                                <div class="section separator"></div>

                                <div class="section">
                                    <p>🙇‍♀️ 对于此次未能准时送达的情况，我们深表歉意。感谢您的理解与支持。</p>
                                </div>

                                <div class="section support">
                                    <h3>需要帮助？ | Need Help?</h3>
                                    <p>✉️ 邮箱: 3555844679@qq.com</p>
                                    <p>🕙 服务群: 1021471813</p>
                                </div>

                                <div class="section">
                                    <img src="cid:footer_image" alt="Service Support" class="footer-image">
                                </div>
                            </div>
                            <div class="footer">
                                <p>LeafAuto Team | {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())} (GMT+8)</p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """

            html_part = MIMEText(html_content, 'html', 'utf-8')
            message.attach(html_part)

            try:
                footer_image_path = get_resource_path('resources/img/page1/email.png')
                with open(footer_image_path, 'rb') as f:
                    footer_image = MIMEImage(f.read())
                    footer_image.add_header('Content-ID', '<footer_image>')
                    message.attach(footer_image)
            except Exception as e:
                log_print(f"[AutoInfo] Failed to attach footer image: {str(e)}")

            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(username, password)
                server.sendmail(sender_email, [receiver_email], message.as_string())
                self.last_email_time = current_time
                log("INFO", f"任务错误邮件发送成功: {task['sender']} -> {task['name']}")
                log_print(f"[AutoInfo] Task error email sent successfully: {task['sender']} -> {task['name']}")

        except Exception as e:
            log_print(f"[AutoInfo] Failed to send error email: {str(e)}")
